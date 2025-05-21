import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook
import os

rh_thickness_csv = "../data/rh_ct_fs_PPMI.csv"
lh_thickness_csv = "../data/lh_ct_fs_PPMI.csv"
rh_area_csv = "../data/rh_sa_fs_PPMI.csv"
lh_area_csv = "../data/lh_sa_fs_PPMI.csv"
cortical_md_csv = "../data/cortical_md_results.csv"
subcortical_csv = "../data/fs_asegstats_PPMI.csv"
demographics_csv = "../data/ppmi_mri_matched_data_karl.csv"

def main():

    csv_list = [rh_thickness_csv, lh_thickness_csv, rh_area_csv, lh_area_csv, cortical_md_csv, subcortical_csv]
    wb_list = ["../data/rh_thickness_demographics.xlsx", "../data/lh_thickness_demographics.xlsx", "../data/rh_area_demographics.xlsx", "../data/lh_area_demographics.xlsx", "../data/cortical_md_demographics.xlsx", "../data/subcortical_demographics.xlsx"]

    csv_wb_list = zip(csv_list, wb_list)

    for csv, wb in csv_wb_list:
        sessions = ["All", "ses-BL", "ses-V04", "ses-V06", "ses-V10"]
        for session in sessions:
            make_demographic_sheet(wb, csv, session)

    return

def make_demographic_sheet(excel_wb, csv, session):
    # Obviously we will need these
    age_dict, sex_dict, educ_dict, diagnosis_dict, updrs_dict, moca_dict = make_demographic_dicts()

    # A bit tricky here... we want demographic stats for before the filtering for the MLM is done, as well as after
    # So we can see how many subjects are lost and how this impacts the demographics
    # Or maybe this ends up not being too hard
    # General strategy will be to look at the dataframe before any filtering
    # Get subject and sessions of interest from there, and calculate demographics using the demographic dictionaries
    # And then look at the dataframe after filtering and see how the demographics change
    df = pd.read_csv(csv)
    subject_col = df.columns[0]
    session_col = df.columns[1]

    # Welcome to the ugliest code I've ever written! Enjoy your stay :^)

    average_age_overall_before, number_male_overall_before, number_female_overall_before, average_educ_overall_before, number_spd_overall_before, number_prod_overall_before, number_hc_overall_before, average_updrs_overall_before, average_moca_overall_before,\
    average_age_spd_before, number_male_spd_before, number_female_spd_before, average_educ_spd_before, average_updrs_spd_before, average_moca_spd_before, \
    average_age_prod_before, number_male_prod_before, number_female_prod_before, average_educ_prod_before, average_updrs_prod_before, average_moca_prod_before, \
    average_age_hc_before, number_male_hc_before, number_female_hc_before, average_educ_hc_before, average_updrs_hc_before, average_moca_hc_before = get_demographics_before(df, session, subject_col, session_col, age_dict, sex_dict, educ_dict, diagnosis_dict, updrs_dict, moca_dict)

    average_age_overall_after, number_male_overall_after, number_female_overall_after, average_educ_overall_after, number_spd_overall_after, number_hc_overall_after, number_prod_overall_after, \
    average_updrs_overall_after, average_updrs_hc_after, average_updrs_prod_after, average_updrs_spd_after, average_moca_overall_after, average_moca_hc_after, average_moca_prod_after, average_moca_spd_after, \
    average_age_spd_after, number_male_spd_after, number_female_spd_after, average_educ_spd_after, \
    average_age_prod_after, number_male_prod_after, number_female_prod_after, average_educ_prod_after, \
    average_age_hc_after, number_male_hc_after, number_female_hc_after, average_educ_hc_after = get_demographics_after_filtering(df, session, subject_col, session_col, age_dict, sex_dict, educ_dict, diagnosis_dict, updrs_dict, moca_dict)

    average_updrs_age, number_updrs_male, number_updrs_female, average_updrs_educ, average_updrs_updrs, average_updrs_moca = get_demographics_after_updrs_filtering(df, session, subject_col, session_col, age_dict, sex_dict, educ_dict, diagnosis_dict, updrs_dict, moca_dict)

    average_moca_age, number_moca_male, number_moca_female, average_moca_educ, average_moca_moca, average_moca_updrs = get_demographics_after_moca_filtering(df, session, subject_col, session_col, age_dict, sex_dict, educ_dict, diagnosis_dict, moca_dict, updrs_dict)

    data_dict = {"Filtering": ['None', 'General', 'UPDRS', 'MoCA'], \
                "Average Overall Age (years)": [average_age_overall_before, average_age_overall_after, 'N/A', 'N/A'], 
                "Number Males Overall": [number_male_overall_before, number_male_overall_after, 'N/A', 'N/A'], 
                "Number Females Overall": [number_female_overall_before, number_female_overall_after, 'N/A', 'N/A'], 
                "Average Education (years)": [average_educ_overall_before, average_educ_overall_after, 'N/A', 'N/A'], 
                "Number sPD Overall": [number_spd_overall_before, number_spd_overall_after, 'N/A', 'N/A'], 
                "Number Prod Overall": [number_prod_overall_before, number_prod_overall_after, 'N/A', 'N/A'], 
                "Number HC Overall": [number_hc_overall_before, number_hc_overall_after, 'N/A', 'N/A'], 
                "Average UPDRS": [average_updrs_overall_before, average_updrs_overall_after, 'N/A', 'N/A'],
                "Average MoCA": [average_moca_overall_before, average_moca_overall_after, 'N/A', 'N/A'],
                "Average Age (years, sPD)": [average_age_spd_before, average_age_spd_after, average_updrs_age, average_moca_age],
                "Number Males (sPD)": [number_male_spd_before, number_male_spd_after, number_updrs_male, number_moca_male],
                "Number Females (sPD)": [number_female_spd_before, number_female_spd_after, number_updrs_female, number_moca_female],
                "Average Education (sPD)": [average_educ_spd_before, average_educ_spd_after, average_updrs_educ, average_moca_educ],
                "Average UPDRS (sPD)": [average_updrs_spd_before, average_updrs_spd_after, average_updrs_updrs, average_moca_updrs],
                "Average MoCA (sPD)": [average_moca_spd_before, average_moca_spd_after, average_updrs_moca, average_moca_moca],
                "Average Age (Prod)": [average_age_prod_before, average_age_prod_after, 'N/A', 'N/A'],
                "Number Male (Prod)": [number_male_prod_before, number_male_prod_after, 'N/A', 'N/A'],
                "Number Female (Prod)": [number_female_prod_before, number_female_prod_after, 'N/A', 'N/A'],
                "Average Educ (Prod)": [average_educ_prod_before, average_educ_prod_after, 'N/A', 'N/A'],
                "Average UPDRS (Prod)": [average_updrs_prod_before, average_updrs_prod_after, 'N/A', 'N/A'],
                "Average MoCA (Prod)": [average_moca_prod_before, average_moca_prod_after, 'N/A', 'N/A'],
                "Average Age (HC)": [average_age_hc_before, average_age_hc_after, 'N/A', 'N/A'],
                "Number Male (HC)": [number_male_hc_before, number_male_hc_after, 'N/A', 'N/A'],
                "Number Female (HC)": [number_female_hc_before, number_female_hc_after, 'N/A', 'N/A'],
                "Average Educ (HC)": [average_educ_hc_before, average_educ_hc_after, 'N/A', 'N/A'],
                "Average UPDRS (HC)": [average_updrs_hc_before, average_updrs_hc_after, 'N/A', 'N/A'],
                "Average MoCA (HC)": [average_moca_hc_before, average_moca_hc_after, 'N/A', 'N/A']
                }
    
    data_df = pd.DataFrame(data_dict)

    if os.path.isfile(excel_wb):
        results_wb = load_workbook(excel_wb)
    else:
        results_wb = create_excel_wb(excel_wb)

    if session in results_wb.sheetnames:
        with pd.ExcelWriter(excel_wb, mode='a', if_sheet_exists='replace') as writer:
            data_df.to_excel(writer, sheet_name=session)
    else:
        if os.path.isfile(excel_wb):
            with pd.ExcelWriter(excel_wb, mode='a') as writer:
                data_df.to_excel(writer, sheet_name=session)
        else:
            with pd.ExcelWriter(excel_wb, mode='w') as writer:
                data_df.to_excel(writer, sheet_name=session)
    return

def get_demographics_after_filtering(df, ses, subject_column, session_column, age_dict, sex_dict, educ_dict, diagnosis_dict, updrs_dict, moca_dict):
    age_values = []
    sex_values = []
    educ_values = []
    diagnosis_values = []
    updrs_values = []
    moca_values = []

    age_spd_values = []
    sex_spd_values = []
    educ_spd_values = []
    updrs_spd_values = []
    moca_spd_values = []

    age_prod_values = []
    sex_prod_values = []
    educ_prod_values = []
    updrs_prod_values = []
    moca_prod_values = []

    age_hc_values = []
    sex_hc_values = []
    educ_hc_values = []
    updrs_hc_values = []
    moca_hc_values = []

    for index, row in df.iterrows():
        subject = row[subject_column]
        session = row[session_column]
        if ses == "ses-BL":
            if session != "ses-BL":
                continue
        elif ses == "ses-V04":
            if session != "ses-V04":
                continue
        elif ses == "ses-V06":
            if session != "ses-V06":
                continue
        elif ses == "ses-V10":
            if session != "ses-V10":
                continue
        if session == "ses-V08": 
            continue
        try:
            age = age_dict[(subject, session)]
        except KeyError:
            continue
        sex = sex_dict[(subject, session)]
        education = educ_dict[(subject, session)]

        diagnosis_values.append(diagnosis_dict[(subject, session)])
        age_values.append(age)
        sex_values.append(sex)
        educ_values.append(education)
        try:
            updrs_values.append(updrs_dict[(subject, session)])
        except KeyError:
            pass
        try:
            moca_values.append(moca_dict[(subject, session)])
        except KeyError:
            pass
        if diagnosis_dict[(subject, session)] == "ctrl":
            age_hc_values.append(age_dict[(subject, session)])
            sex_hc_values.append(sex_dict[(subject, session)])
            educ_hc_values.append(educ_dict[(subject, session)])
            try:
                updrs_hc_values.append(updrs_dict[(subject, session)])
            except KeyError:
                pass
            try:
                moca_hc_values.append(moca_dict[(subject, session)])
            except KeyError:
                pass
        elif diagnosis_dict[(subject, session)] == "prod":
            age_prod_values.append(age_dict[(subject, session)])
            sex_prod_values.append(sex_dict[(subject, session)])
            educ_prod_values.append(educ_dict[(subject, session)])
            try:
                updrs_prod_values.append(updrs_dict[(subject, session)])
            except KeyError:
                pass
            try:
                moca_prod_values.append(moca_dict[(subject, session)])
            except KeyError:
                pass
        elif diagnosis_dict[(subject, session)] == "park":
            age_spd_values.append(age_dict[(subject, session)])
            sex_spd_values.append(sex_dict[(subject, session)])
            educ_spd_values.append(educ_dict[(subject, session)])
            try:
                updrs_spd_values.append(updrs_dict[(subject, session)])
            except KeyError:
                pass
            try:
                moca_spd_values.append(moca_dict[(subject, session)])
            except KeyError:
                pass 
    average_age_overall = sum(age_values)/len(age_values)
    number_male_overall = sex_values.count('male')
    number_female_overall = sex_values.count('female')
    average_educ_overall = sum(educ_values)/len(educ_values)
    number_spd_overall = diagnosis_values.count('park')
    number_hc_overall = diagnosis_values.count('ctrl')
    number_prod_overall = diagnosis_values.count('prod')
    average_updrs_overall = sum(updrs_values)/len(updrs_values)
    average_moca_overall = sum(moca_values)/len(moca_values)

    average_age_spd = sum(age_spd_values)/len(age_spd_values)
    number_male_spd = sex_spd_values.count('male')
    number_female_spd = sex_spd_values.count('female')
    average_educ_spd = sum(educ_spd_values)/len(educ_spd_values)
    average_updrs_spd = sum(updrs_spd_values)/len(updrs_spd_values)
    average_moca_spd = sum(moca_spd_values)/len(moca_spd_values)

    average_age_prod = sum(age_prod_values)/len(age_prod_values)
    number_male_prod = sex_prod_values.count('male')
    number_female_prod = sex_prod_values.count('female')
    average_educ_prod = sum(educ_prod_values)/len(educ_prod_values)
    average_updrs_prod = sum(updrs_prod_values)/len(updrs_prod_values)
    average_moca_prod = sum(moca_prod_values)/len(moca_prod_values)

    average_age_hc = sum(age_hc_values)/len(age_hc_values)
    number_male_hc = sex_hc_values.count('male')
    number_female_hc = sex_hc_values.count('female')
    average_educ_hc = sum(educ_hc_values)/len(educ_hc_values)
    average_updrs_hc = sum(updrs_hc_values)/len(updrs_hc_values)
    average_moca_hc = sum(moca_hc_values)/len(moca_hc_values)


    return average_age_overall, number_male_overall, number_female_overall, average_educ_overall, number_spd_overall, number_hc_overall, number_prod_overall, \
    average_updrs_overall, average_updrs_hc, average_updrs_prod, average_updrs_spd, average_moca_overall, average_moca_hc, average_moca_prod, average_moca_spd, \
    average_age_spd, number_male_spd, number_female_spd, average_educ_spd, \
    average_age_prod, number_male_prod, number_female_prod, average_educ_prod, \
    average_age_hc, number_male_hc, number_female_hc, average_educ_hc

def get_demographics_after_updrs_filtering(df, ses, subject_column, session_column, age_dict, sex_dict, educ_dict, diagnosis_dict, updrs_dict, moca_dict):
    age_values = []
    sex_values = []
    educ_values = []
    updrs_values = []
    moca_values = []

    for index, row in df.iterrows():
        subject = row[subject_column]
        session = row[session_column]
        if ses == "ses-BL":
            if session != "ses-BL":
                continue
        elif ses == "ses-V04":
            if session != "ses-V04":
                continue
        elif ses == "ses-V06":
            if session != "ses-V06":
                continue
        elif ses == "ses-V10":
            if session != "ses-V10":
                continue
        if session == "ses-V08":
            continue
        try:
            age = age_dict[(subject, session)]
        except KeyError:
            continue
        sex = sex_dict[(subject, session)]
        educ = educ_dict[(subject, session)]
        try:
            updrs = updrs_dict[(subject, session)]
        except KeyError:
            continue

        diagnosis = diagnosis_dict[(subject, session)]

        if diagnosis == 'park':
            age_values.append(age)
            sex_values.append(sex)
            educ_values.append(educ)
            updrs_values.append(updrs)
            try:
                moca_values.append(moca_dict[(subject, session)])
            except KeyError:
                pass

    average_age = sum(age_values)/len(age_values)
    number_male = sex_values.count('male')
    number_female = sex_values.count('female')
    average_educ = sum(educ_values)/len(educ_values)
    average_updrs = sum(updrs_values)/len(updrs_values)
    average_moca = sum(moca_values)/len(moca_values)

    return average_age, number_male, number_female, average_educ, average_updrs, average_moca

def get_demographics_after_moca_filtering(df, ses, subject_column, session_column, age_dict, sex_dict, educ_dict, diagnosis_dict, moca_dict, updrs_dict):
    age_values = []
    sex_values = []
    educ_values = []
    moca_values = []
    updrs_values = []

    for index, row in df.iterrows():
        subject = row[subject_column]
        session = row[session_column]
        if ses == "ses-BL":
            if session != "ses-BL":
                continue
        elif ses == "ses-V04":
            if session != "ses-V04":
                continue
        elif ses == "ses-V06":
            if session != "ses-V06":
                continue
        elif ses == "ses-V10":
            if session != "ses-V10":
                continue
        if session == "ses-V08":
            continue
        try:
            age = age_dict[(subject, session)]
        except KeyError:
            continue
        sex = sex_dict[(subject, session)]
        educ = educ_dict[(subject, session)]
        try:
            moca = moca_dict[(subject, session)]
        except KeyError:
            continue

        diagnosis = diagnosis_dict[(subject, session)]

        if diagnosis == 'park':
            age_values.append(age)
            sex_values.append(sex)
            educ_values.append(educ)
            moca_values.append(moca)
            try:
                updrs_values.append(updrs_dict[(subject, session)])
            except KeyError:
                pass

    average_age = sum(age_values)/len(age_values)
    number_male = sex_values.count('male')
    number_female = sex_values.count('female')
    average_educ = sum(educ_values)/len(educ_values)
    average_moca = sum(moca_values)/len(moca_values)
    average_updrs = sum(updrs_values)/len(updrs_values)

    return average_age, number_male, number_female, average_educ, average_moca, average_updrs

def get_demographics_before(df, ses, subject_column, session_column, age_dict, sex_dict, educ_dict, diagnosis_dict, updrs_dict, moca_dict):
    age_values = []
    sex_values = []
    educ_values = []
    diagnosis_values = []
    updrs_values = []
    moca_values = []

    age_spd_values = []
    sex_spd_values = []
    educ_spd_values = []
    updrs_spd_values = []
    moca_spd_values = []

    age_prod_values = []
    sex_prod_values = []
    educ_prod_values = []
    updrs_prod_values = []
    moca_prod_values = []

    age_hc_values = []
    sex_hc_values = []
    educ_hc_values = []
    updrs_hc_values = []
    moca_hc_values = []

    for index, row in df.iterrows():
        subject = row[subject_column]
        session = row[session_column]
        if ses == "ses-BL":
            if session != "ses-BL":
                continue
        elif ses == "ses-V04":
            if session != "ses-V04":
                continue
        elif ses == "ses-V06":
            if session != "ses-V06":
                continue
        elif ses == "ses-V10":
            if session != "ses-V10":
                continue
        try:
            age_values.append(age_dict[(subject, session)])
        except KeyError:
            pass
        try:
            sex_values.append(sex_dict[(subject, session)])
        except KeyError:
            pass
        try:
            educ_values.append(educ_dict[(subject, session)])
        except KeyError:
            pass
        try:
            diagnosis_values.append(diagnosis_dict[(subject, session)])
        except KeyError:
            pass
        try:
            updrs_values.append(updrs_dict[(subject, session)])
        except KeyError:
            pass
        try:
            moca_values.append(moca_dict[(subject, session)])
        except KeyError:
            pass
        try:
            if diagnosis_dict[(subject, session)] == "ctrl":
                age_hc_values.append(age_dict[(subject, session)])
                sex_hc_values.append(sex_dict[(subject, session)])
                educ_hc_values.append(educ_dict[(subject, session)])
                updrs_hc_values.append(updrs_dict[(subject, session)])
                moca_hc_values.append(moca_dict[(subject, session)])
            elif diagnosis_dict[(subject, session)] == "prod":
                age_prod_values.append(age_dict[(subject, session)])
                sex_prod_values.append(sex_dict[(subject, session)])
                educ_prod_values.append(educ_dict[(subject, session)])
                updrs_prod_values.append(updrs_dict[(subject, session)])
                moca_prod_values.append(moca_dict[(subject, session)])
            elif diagnosis_dict[(subject, session)] == "park":
                age_spd_values.append(age_dict[(subject, session)])
                sex_spd_values.append(sex_dict[(subject, session)])
                educ_spd_values.append(educ_dict[(subject, session)])
                updrs_spd_values.append(updrs_dict[(subject, session)])
                moca_spd_values.append(moca_dict[(subject, session)])
        except KeyError:
            pass

    average_age_overall = sum(age_values)/len(age_values)
    number_male_overall = sex_values.count('male')
    number_female_overall = sex_values.count('female')
    average_educ_overall = sum(educ_values)/len(educ_values)
    number_spd_overall = diagnosis_values.count('park')
    number_hc_overall = diagnosis_values.count('ctrl')
    number_prod_overall = diagnosis_values.count('prod')
    average_updrs_overall = sum(updrs_values)/len(updrs_values)
    average_moca_overall = sum(moca_values)/len(moca_values)

    average_age_spd = sum(age_spd_values)/len(age_spd_values)
    number_male_spd = sex_spd_values.count('male')
    number_female_spd = sex_spd_values.count('female')
    average_educ_spd = sum(educ_spd_values)/len(educ_spd_values)
    average_updrs_spd = sum(updrs_spd_values)/len(updrs_spd_values)
    average_moca_spd = sum(moca_spd_values)/len(moca_spd_values)

    average_age_prod = sum(age_prod_values)/len(age_prod_values)
    number_male_prod = sex_prod_values.count('male')
    number_female_prod = sex_prod_values.count('female')
    average_educ_prod = sum(educ_prod_values)/len(educ_prod_values)
    average_updrs_prod = sum(updrs_prod_values)/len(updrs_prod_values)
    average_moca_prod = sum(moca_prod_values)/len(moca_prod_values)

    average_age_hc = sum(age_hc_values)/len(age_hc_values)
    number_male_hc = sex_hc_values.count('male')
    number_female_hc = sex_hc_values.count('female')
    average_educ_hc = sum(educ_hc_values)/len(educ_hc_values)
    average_updrs_hc = sum(updrs_hc_values)/len(updrs_hc_values)
    average_moca_hc = sum(moca_hc_values)/len(moca_hc_values)

    # Father forgive me this is gonna be ugly
    return average_age_overall, number_male_overall, number_female_overall, average_educ_overall, number_spd_overall, number_prod_overall, number_hc_overall, average_updrs_overall, average_moca_overall,\
    average_age_spd, number_male_spd, number_female_spd, average_educ_spd, average_updrs_spd, average_moca_spd, \
    average_age_prod, number_male_prod, number_female_prod, average_educ_prod, average_updrs_prod, average_moca_prod, \
    average_age_hc, number_male_hc, number_female_hc, average_educ_hc, average_updrs_hc, average_moca_hc

def make_dictionary(dataframe):

    subject_dict = {}

    df_cols = list(dataframe.columns)
    subject_col = df_cols[0]
    session_col = df_cols[1]
    region_cols = df_cols[2:]

    for index, row in dataframe.iterrows():
        subject = row[subject_col]
        session = row[session_col]
        region_results_dict = {}
        for region in region_cols:
            # We don't want this region
            if "rh_temporalpole" in region:
                continue
            region_results_dict[region] = row[region]
        subject_dict[(subject, session)] = region_results_dict
    return subject_dict

def make_demographic_dicts():
    age_dict = {}
    sex_dict = {}
    educ_dict = {}
    diagnosis_dict = {}
    updrs_dict = {}
    moca_dict = {}

    demographic_df = pd.read_csv(demographics_csv)

    for index, row in demographic_df.iterrows():
        subject = f"sub-{row['participant_id']}"
        session = f"ses-{row['visit']}"
        sex = row['sex']
        age = row['age']
        education = row['education']
        diagnosis = row['diagnosis']
        updrs = row['updrs3']
        moca = row['moca']

        # Check for missing values, skip if there are any since they are necessary for MLM downstream
        if not np.isnan(age) and not np.isnan(education) and sex != np.nan and diagnosis != np.nan: # Accounting for potential missing data
            age_dict[(subject, session)] = age
            sex_dict[(subject, session)] = sex
            educ_dict[(subject, session)] = education
            diagnosis_dict[(subject, session)] = diagnosis

        if not np.isnan(updrs):
            updrs_dict[(subject, session)] = updrs

        if not np.isnan(moca):
            moca_dict[(subject, session)] = moca

    return age_dict, sex_dict, educ_dict, diagnosis_dict, updrs_dict, moca_dict

def create_excel_wb(filepath):
    wb = openpyxl.Workbook()
    wb.save(filepath)
    return wb

if __name__ == "__main__":
    main()