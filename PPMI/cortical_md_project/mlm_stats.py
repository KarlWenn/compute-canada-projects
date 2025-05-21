import pandas as pd
import numpy as np
import statsmodels.formula.api as smf
import openpyxl
from openpyxl import load_workbook
import os

rh_thickness_csv = "../data/rh_ct_fs_PPMI.csv"
lh_thickness_csv = "../data/lh_ct_fs_PPMI.csv"
rh_area_csv = "../data/rh_sa_fs_PPMI.csv"
lh_area_csv = "../data/lh_sa_fs_PPMI.csv"
cortical_md_csv = "../data/cortical_md_results.csv"
subcortical_csv = "../data/fs_asegstats_PPMI.csv"
demographics_csv = "../data/ppmi_mri_matched_data_karl.csv"

def main():
    
    # Get our demographic dictionaries
    age_dict, sex_dict, educ_dict, diagnosis_dict, updrs_dict, moca_dict = make_demographic_dicts()

    # Run MLM on cortical MD measures
    cortical_md_df = pd.read_csv(cortical_md_csv)
    # Results will be stored in a separate Excel file for each brain health measure thing we're looking at
    cortical_md_wb_path = "../data/BL_newest_cortical_md_results.xlsx"
    cortical_md_updrs_wb_path = "../data/BL_new_UPDRS_cortical_md_results.xlsx"
    cortical_md_moca_wb_path= "../data/BL_new_MoCA_cortical_md_results.xlsx"
    cortical_md_dict = make_dictionary(cortical_md_df)
    cortical_md_cols = list(cortical_md_df.columns)[2:] # Need to get column names to iterate through them, don't want subject or session column which are always first two
    run_updrs_mlm_on_dict(cortical_md_updrs_wb_path, cortical_md_dict, cortical_md_cols, age_dict, sex_dict, educ_dict, diagnosis_dict, updrs_dict)
    run_moca_mlm_on_dict(cortical_md_moca_wb_path, cortical_md_dict, cortical_md_cols, age_dict, sex_dict, educ_dict, diagnosis_dict, moca_dict)
    run_mlm_on_dict(cortical_md_wb_path, cortical_md_dict, cortical_md_cols, age_dict, sex_dict, educ_dict, diagnosis_dict) # Using column names to iterate, see function

    # From hereon, it's the same strategy, just repeated
    # Run MLM on CT measures
    rh_thickness_df = pd.read_csv(rh_thickness_csv)
    lh_thickness_df = pd.read_csv(lh_thickness_csv)

    rh_thickness_dict = make_dictionary(rh_thickness_df)
    lh_thickness_dict = make_dictionary(lh_thickness_df)

    rh_thickness_cols = list(rh_thickness_df.columns)[2:]
    lh_thickness_cols = list(lh_thickness_df.columns)[2:]

    rh_thickness_wb_path = "../data/BL_newest_rh_thickness_results.xlsx"
    lh_thickness_wb_path = "../data/BL_newest_lh_thickness_results.xlsx"

    rh_thickness_updrs_wb_path = "../data/BL_new_UPDRS_rh_thickness_results.xlsx"
    lh_thickness_updrs_wb_path = "../data/BL_new_UPDRS_lh_thickness_results.xlsx"

    rh_thickness_moca_wb_path = "../data/BL_new_MoCA_rh_thickness_results.xlsx"
    lh_thickness_moca_wb_path = "../data/BL_new_MoCA_lh_thickness_results.xlsx"

    run_updrs_mlm_on_dict(rh_thickness_updrs_wb_path, rh_thickness_dict, rh_thickness_cols, age_dict, sex_dict, educ_dict, diagnosis_dict, updrs_dict)
    run_updrs_mlm_on_dict(lh_thickness_updrs_wb_path, lh_thickness_dict, lh_thickness_cols, age_dict, sex_dict, educ_dict, diagnosis_dict, updrs_dict)

    run_moca_mlm_on_dict(rh_thickness_moca_wb_path, rh_thickness_dict, rh_thickness_cols, age_dict, sex_dict, educ_dict, diagnosis_dict, moca_dict)
    run_moca_mlm_on_dict(lh_thickness_moca_wb_path, lh_thickness_dict, lh_thickness_cols, age_dict, sex_dict, educ_dict, diagnosis_dict, moca_dict)

    run_mlm_on_dict(rh_thickness_wb_path, rh_thickness_dict, rh_thickness_cols, age_dict, sex_dict, educ_dict, diagnosis_dict)
    run_mlm_on_dict(lh_thickness_wb_path, lh_thickness_dict, lh_thickness_cols, age_dict, sex_dict, educ_dict, diagnosis_dict)

    # Run MLM on SA measures
    rh_sa_df = pd.read_csv(rh_area_csv)
    lh_sa_df = pd.read_csv(lh_area_csv)

    rh_sa_dict = make_dictionary(rh_sa_df)
    lh_sa_dict = make_dictionary(lh_sa_df)

    rh_sa_cols = list(rh_sa_df.columns)[2:]
    lh_sa_cols = list(lh_sa_df.columns)[2:]

    rh_sa_wb_path = "../data/BL_newest_rh_sa_results.xlsx"
    lh_sa_wb_path = "../data/BL_newest_lh_sa_results.xlsx"

    rh_sa_updrs_wb_path = "../data/BL_new_UPDRS_rh_area_results.xlsx"
    lh_sa_updrs_wb_path = "../data/BL_new_UPDRS_lh_area_results.xlsx"

    rh_sa_moca_wb_path = "../data/BL_new_MoCA_rh_area_results.xlsx"
    lh_sa_moca_wb_path = "../data/BL_new_MoCA_lh_area_results.xlsx"

    run_updrs_mlm_on_dict(rh_sa_updrs_wb_path, rh_sa_dict, rh_sa_cols, age_dict, sex_dict, educ_dict, diagnosis_dict, updrs_dict)
    run_updrs_mlm_on_dict(lh_sa_updrs_wb_path, lh_sa_dict, lh_sa_cols, age_dict, sex_dict, educ_dict, diagnosis_dict, updrs_dict)

    run_moca_mlm_on_dict(rh_sa_moca_wb_path, rh_sa_dict, rh_sa_cols, age_dict, sex_dict, educ_dict, diagnosis_dict, moca_dict)
    run_moca_mlm_on_dict(lh_sa_moca_wb_path, lh_sa_dict, lh_sa_cols, age_dict, sex_dict, educ_dict, diagnosis_dict, moca_dict)

    run_mlm_on_dict(rh_sa_wb_path, rh_sa_dict, rh_sa_cols, age_dict, sex_dict, educ_dict, diagnosis_dict)
    run_mlm_on_dict(lh_sa_wb_path, lh_sa_dict, lh_sa_cols, age_dict, sex_dict, educ_dict, diagnosis_dict)

    # Run MLM on subcortical measures
    subcortical_df = pd.read_csv(subcortical_csv)

    subcortical_dict = make_dictionary(subcortical_df)

    subcortical_cols = list(subcortical_df.columns)[2:]

    subcortical_wb_path = "../data/BL_newest_subcortical_results.xlsx"

    subcortical_updrs_wb_path = "../data/BL_new_UPDRS_subcortical_results.xlsx"
    subcortical_moca_wb_path = "../data/BL_new_MoCA_subcortical_results.xlsx"

    run_updrs_mlm_on_dict(subcortical_updrs_wb_path, subcortical_dict, subcortical_cols, age_dict, sex_dict, educ_dict, diagnosis_dict, updrs_dict)
    run_moca_mlm_on_dict(subcortical_moca_wb_path, subcortical_dict, subcortical_cols, age_dict, sex_dict, educ_dict, diagnosis_dict, moca_dict)
    run_mlm_on_dict(subcortical_wb_path, subcortical_dict, subcortical_cols, age_dict, sex_dict, educ_dict, diagnosis_dict)

    return

def make_demographic_dicts():
    age_dict = {}
    sex_dict = {}
    educ_dict = {}
    diagnosis_dict = {}
    updrs_dict = {}
    moca_dict = {}

    demographic_df = pd.read_csv(demographics_csv)

    for index, row in demographic_df.iterrows():
        subject = f"sub-{row['participant_id']}"
        session = f"ses-{row['visit']}"
        sex = row['sex']
        age = row['age']
        education = row['education']
        diagnosis = row['diagnosis']
        updrs = row['updrs3']
        moca = row['moca']

        # Check for missing values, skip if there are any since they are necessary for MLM downstream
        if not np.isnan(age) and not np.isnan(education) and sex != np.nan and diagnosis != np.nan: # Accounting for potential missing data
            age_dict[(subject, session)] = age
            sex_dict[(subject, session)] = sex
            educ_dict[(subject, session)] = education
            diagnosis_dict[(subject, session)] = diagnosis

        if not np.isnan(updrs):
            updrs_dict[(subject, session)] = updrs

        if not np.isnan(moca):
            moca_dict[(subject, session)] = moca

    return age_dict, sex_dict, educ_dict, diagnosis_dict, updrs_dict, moca_dict

def run_updrs_mlm_on_dict(excel_wb_path, region_dict, column_names, age_dict, sex_dict, educ_dict, diagnosis_dict, updrs_dict):
    for region in column_names:
        subject_values = []
        region_values = []
        group_values = []
        visit_values = []
        age_values = []
        sex_values = []
        educ_values = []
        updrs_values = []

        if "rh_temporalpole" in region or "Left-WM-hypointensities" in region or "Right-WM-hypointensities" in region or "Left-non-WM-hypointensities" in region or "Right-non-WM-hypointensities" in region:
            continue
        for (subject, session) in region_dict:
            # Skip session 8?
            if session != "ses-BL":
                continue
            if session == "ses-V08":
                continue
            # To make visit values continuous
            elif session == "ses-BL":
                visit = 1
            elif session == "ses-V04":
                visit = 2
            elif session == "ses-V06":
                visit = 3
            elif session == "ses-V10":
                visit = 5
            region_value = region_dict[(subject, session)][region] # Skip subjects missing the value for this region
            if np.isnan(region_value):
                continue
            try:
                age = age_dict[(subject, session)]
            except KeyError:
                #print(f"Skipping {subject} {session}, genetic or prodromal PD")
                continue
            sex = sex_dict[(subject, session)]
            education = educ_dict[(subject, session)]
            try:
                updrs = updrs_dict[(subject, session)]
            except KeyError: # No UPDRS for this subject
                continue

            # Make education values categorical
            if education <= 12:
                educ_level = 1
            elif education <= 16:
                educ_level = 2
            elif education > 16:
                educ_level = 3

            diagnosis = diagnosis_dict[(subject, session)]
            
            if diagnosis == 'park':
                subject_values.append(subject)
                visit_values.append(visit)
                age_values.append(age)
                sex_values.append(sex)
                educ_values.append(educ_level)
                group_values.append(diagnosis)
                region_values.append(region_value)
                updrs_values.append(updrs)
            else: 
                continue
        print(region)
        updrs_result = run_updrs_mlm(region_values, visit_values, subject_values, age_values, sex_values, educ_values, updrs_values)
        # Have to update the Excel workbook for the region
        write_updrs_excel_sheets(updrs_result, region, excel_wb_path)
    return 

def run_moca_mlm_on_dict(excel_wb_path, region_dict, column_names, age_dict, sex_dict, educ_dict, diagnosis_dict, moca_dict):
    for region in column_names:
        subject_values = []
        region_values = []
        group_values = []
        visit_values = []
        age_values = []
        sex_values = []
        educ_values = []
        moca_values = []

        if "rh_temporalpole" in region or "Left-WM-hypointensities" in region or "Right-WM-hypointensities" in region or "Left-non-WM-hypointensities" in region or "Right-non-WM-hypointensities" in region:
            continue
        for (subject, session) in region_dict:
            # Skip session 8?
            if session != "ses-BL":
                continue
            if session == "ses-V08":
                continue
            # To make visit values continuous
            elif session == "ses-BL":
                visit = 1
            elif session == "ses-V04":
                visit = 2
            elif session == "ses-V06":
                visit = 3
            elif session == "ses-V10":
                visit = 5
            region_value = region_dict[(subject, session)][region] # Skip subjects missing the value for this region
            if np.isnan(region_value):
                continue
            try:
                age = age_dict[(subject, session)]
            except KeyError:
                #print(f"Skipping {subject} {session}, genetic or prodromal PD")
                continue
            sex = sex_dict[(subject, session)]
            education = educ_dict[(subject, session)]
            try:
                moca = moca_dict[(subject, session)]
            except KeyError: # No moca for this subject
                continue

            # Make education values categorical
            if education <= 12:
                educ_level = 1
            elif education <= 16:
                educ_level = 2
            elif education > 16:
                educ_level = 3

            diagnosis = diagnosis_dict[(subject, session)]
            
            if diagnosis == 'park':
                subject_values.append(subject)
                visit_values.append(visit)
                age_values.append(age)
                sex_values.append(sex)
                educ_values.append(educ_level)
                group_values.append(diagnosis)
                region_values.append(region_value)
                moca_values.append(moca)
            else: 
                continue
        print(region)
        moca_result = run_moca_mlm(region_values, visit_values, subject_values, age_values, sex_values, educ_values, moca_values)
        # Have to update the Excel workbook for the region
        write_moca_excel_sheets(moca_result, region, excel_wb_path)
    return 

def write_moca_excel_sheets(moca_result, region, excel_wb_path):

    sex_beta = moca_result.params[1]
    sex_pvalue = moca_result.pvalues[1]
    moca_beta = moca_result.params[2]
    moca_pvalue = moca_result.pvalues[2]
    age_beta = moca_result.params[3]
    age_pvalue = moca_result.pvalues[3]
    #visit_beta = moca_result.params[4]
    #visit_pvalue = moca_result.pvalues[4]
    educ_beta = moca_result.params[4]
    educ_pvalue = moca_result.pvalues[4]

    result_dict = {
        "Parameter": ['Sex', 'Educ', 'MoCA', 'Age'],
        "Beta": [sex_beta, educ_beta, moca_beta, age_beta],
        "P-Value": [sex_pvalue, educ_pvalue, moca_pvalue, age_pvalue]
    }

    results_df = pd.DataFrame(result_dict)

    if os.path.isfile(excel_wb_path):
        results_wb = load_workbook(excel_wb_path)
    else:
        results_wb = create_excel_wb(excel_wb_path)

    if region in results_wb.sheetnames:
        with pd.ExcelWriter(excel_wb_path, mode='a', if_sheet_exists='replace') as writer:
            results_df.to_excel(writer, sheet_name=region)
    else:
        if os.path.isfile(excel_wb_path):
            with pd.ExcelWriter(excel_wb_path, mode='a') as writer:
                results_df.to_excel(writer, sheet_name=region)
        else:
            with pd.ExcelWriter(excel_wb_path, mode='w') as writer:
                results_df.to_excel(writer, sheet_name=region)

    return

def write_updrs_excel_sheets(updrs_result, region, excel_wb_path):
    
    # More typing ZZZZZ
    sex_beta = updrs_result.params[1]
    sex_pvalue = updrs_result.pvalues[1]
    updrs_beta = updrs_result.params[2]
    updrs_pvalue = updrs_result.pvalues[2]
    age_beta = updrs_result.params[3]
    age_pvalue = updrs_result.pvalues[3]
    #visit_beta = updrs_result.params[4]
    #visit_pvalue = updrs_result.pvalues[4]
    educ_beta = updrs_result.params[4]
    educ_pvalue = updrs_result.pvalues[4]

    result_dict = {
        "Parameter": ['Sex', 'Educ', 'UPDRS', 'Age'],
        "Beta": [sex_beta, educ_beta, updrs_beta, age_beta],
        "P-Value": [sex_pvalue, educ_pvalue, updrs_pvalue, age_pvalue]
    }
    
    results_df = pd.DataFrame(result_dict)

    if os.path.isfile(excel_wb_path):
        results_wb = load_workbook(excel_wb_path)
    else: 
        results_wb = create_excel_wb(excel_wb_path)

    if region in results_wb.sheetnames:
        with pd.ExcelWriter(excel_wb_path, mode='a', if_sheet_exists='replace') as writer:
            results_df.to_excel(writer, sheet_name=region)
    else:
        if os.path.isfile(excel_wb_path):
            with pd.ExcelWriter(excel_wb_path, mode='a') as writer:
                results_df.to_excel(writer, sheet_name=region)
        else:
            with pd.ExcelWriter(excel_wb_path, mode='w') as writer:
                results_df.to_excel(writer, sheet_name=region)

    return

def run_updrs_mlm(region_values, visit_values, subject_values, age_values, sex_values, educ_values, updrs_values):
    data_dict = {"dependent_values": region_values, "visit": visit_values, "subject": subject_values, "age": age_values, "sex": sex_values, "education": educ_values, "updrs": updrs_values}
    df = pd.DataFrame(data_dict)

    # region_values ~ updrs + age + sex + visit + education + (updrs|subj)
    #vc = {'updrs': '0 + updrs'}
    updrs_model = smf.glm("dependent_values ~ updrs + age + sex + education",
                              data = df)
    
    updrs_result = updrs_model.fit()
    print(updrs_result.summary())
    
    return updrs_result

def run_moca_mlm(region_values, visit_values, subject_values, age_values, sex_values, educ_values, moca_values):

    data_dict = {"dependent_values": region_values, "visit": visit_values, "subject": subject_values, "age": age_values, "sex": sex_values, "education": educ_values, "moca": moca_values}
    df = pd.DataFrame(data_dict)
    
    # region_values ~ moca + age + sex + visit + education + (moca|subj)
    #vc = {'moca': '0 + moca'}
    moca_model = smf.glm("dependent_values ~ moca + age + sex + education",
                             data = df)
    
    moca_result = moca_model.fit()
    print(moca_result.summary())

    return moca_result

def run_mlm_on_dict(excel_wb_path, region_dict, column_names, age_dict, sex_dict, educ_dict, diagnosis_dict):
    # Takes in dictionary of form {(subject, session): {region_1: ..., region_2: ..., region_3: ...}, (subject, session): {...}, ...}
    # Runs a MLM for every region across every subject session

    # Run an MLM for every region, on every subject session
    for region in column_names:
        subject_values = []
        region_values = []
        group_values = []
        visit_values = []
        age_values = []
        sex_values = []
        educ_values = []

        if "rh_temporalpole" in region or "Left-WM-hypointensities" in region or "Right-WM-hypointensities" in region or "Left-non-WM-hypointensities" in region or "Right-non-WM-hypointensities" in region:
            continue
        for (subject, session) in region_dict:
            try:
                diagnosis = diagnosis_dict[(subject, session)]
            except KeyError:
                continue

            # Make group values longitudinal
            if diagnosis == 'park':
                group = 1
            elif diagnosis == 'prod':
                continue
            elif diagnosis == 'ctrl':
                group = 0
            if session != "ses-BL":
                continue
            # Skip session 8; not enough subjects, causes MLM to fail
            if session == "ses-V08":
                continue
            # To make visit values continuous
            elif session == "ses-BL":
                visit = 1
            elif session == "ses-V04":
                visit = 2
            elif session == "ses-V06":
                visit = 3
            elif session == "ses-V10":
                visit = 5
            region_value = region_dict[(subject, session)][region] # Skip subjects missing the value for this region
            if np.isnan(region_value):
                continue
            try:
                age = age_dict[(subject, session)]
            except KeyError:
                #print(f"Skipping {subject} {session}, genetic or prodromal PD")
                continue
            subject_values.append(subject)
            visit_values.append(visit)
            sex = sex_dict[(subject, session)]
            education = educ_dict[(subject, session)]

            # Make education values categorical
            if education <= 12:
                educ_level = 1
            elif education <= 16:
                educ_level = 2
            elif education > 16:
                educ_level = 3

            age_values.append(age)
            sex_values.append(sex)
            educ_values.append(educ_level)
            group_values.append(group)
            region_values.append(region_value)
        print(region)
        model_result = run_mlm(region_values, group_values, visit_values, subject_values, age_values, sex_values, educ_values)
        # Have to update the Excel workbook for the region
        write_excel_sheets(model_result, region, excel_wb_path)
    return 

def write_excel_sheets(model_result, region, excel_wb_path):

    # A lot of typing... zzzz....
    sex_beta = model_result.params[1]
    sex_pvalue = model_result.pvalues[1]
    group_beta = model_result.params[2]
    group_pvalue = model_result.pvalues[2]
    #visit_beta = model_result.params[3]
    #visit_pvalue = model_result.pvalues[3]
    #group_x_visit_beta = model_result.params[4]
    #group_x_visit_pvalue = model_result.pvalues[4]
    age_beta = model_result.params[3]
    age_pvalue = model_result.pvalues[3]
    educ_beta = model_result.params[4]
    educ_pvalue = model_result.pvalues[4]

    result_dict = {
        "Parameter": ['Sex', 'Group', 'Age', 'Educ'],
        "Beta": [sex_beta, group_beta, age_beta, educ_beta],
        "P-Value": [sex_pvalue, group_pvalue, age_pvalue, educ_pvalue]
    }

    results_df = pd.DataFrame(result_dict)

    if os.path.isfile(excel_wb_path):
        results_wb = load_workbook(excel_wb_path)
    else:
        results_wb = create_excel_wb(excel_wb_path)

    if region in results_wb.sheetnames:
        with pd.ExcelWriter(excel_wb_path, mode='a', if_sheet_exists='replace') as writer:
            results_df.to_excel(writer, sheet_name=region)
    else:
        if os.path.isfile(excel_wb_path):
            with pd.ExcelWriter(excel_wb_path, mode='a') as writer:
                results_df.to_excel(writer, sheet_name=region)
        else:
            with pd.ExcelWriter(excel_wb_path, mode='w') as writer:
                results_df.to_excel(writer, sheet_name=region)
    return

def make_dictionary(dataframe):

    subject_dict = {}

    df_cols = list(dataframe.columns)
    subject_col = df_cols[0]
    session_col = df_cols[1]
    region_cols = df_cols[2:]

    for index, row in dataframe.iterrows():
        subject = row[subject_col]
        session = row[session_col]
        region_results_dict = {}
        for region in region_cols:
            # We don't want this region
            if "rh_temporalpole" in region:
                continue
            region_results_dict[region] = row[region]
        subject_dict[(subject, session)] = region_results_dict
    return subject_dict

def run_mlm(dependent_values, group_values, visit_values, subject_values, age_values, sex_values, educ_values):
    data_dict = {"dependent_values": dependent_values, "group": group_values, "visit": visit_values, "subject": subject_values, "age": age_values, "sex": sex_values, "education": educ_values}
    df = pd.DataFrame(data_dict)

    #vc = {'visit': '0 + visit'}
    model = smf.glm("dependent_values ~ group + age + sex + education",
                        data=df)
    
    model_result = model.fit()
    print(model_result.summary())
    
    return model_result

def create_excel_wb(filepath):
    wb = openpyxl.Workbook()
    wb.save(filepath)
    return wb

if __name__ == "__main__":
    main()