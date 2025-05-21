from statsmodels.stats.multitest import fdrcorrection
import pandas as pd
import openpyxl

def main():
    subcortical_wb = openpyxl.load_workbook("../data/BL_newest_subcortical_results.xlsx")
    rh_thickness_wb = openpyxl.load_workbook("../data/BL_newest_rh_thickness_results.xlsx")
    lh_thickness_wb = openpyxl.load_workbook("../data/BL_newest_lh_thickness_results.xlsx")
    rh_area_wb = openpyxl.load_workbook("../data/BL_newest_rh_sa_results.xlsx")
    lh_area_wb = openpyxl.load_workbook("../data/BL_newest_lh_sa_results.xlsx")
    cortical_md_wb = openpyxl.load_workbook("../data/BL_newest_cortical_md_results.xlsx")

    updrs_cortical_md_wb = openpyxl.load_workbook("../data/BL_new_UPDRS_cortical_md_results.xlsx")
    updrs_lh_thickness_wb = openpyxl.load_workbook("../data/BL_new_UPDRS_lh_thickness_results.xlsx")
    updrs_rh_thickness_wb = openpyxl.load_workbook("../data/BL_new_UPDRS_rh_thickness_results.xlsx")
    updrs_lh_area_wb = openpyxl.load_workbook("../data/BL_new_UPDRS_lh_area_results.xlsx")
    updrs_rh_area_wb = openpyxl.load_workbook("../data/BL_new_UPDRS_rh_area_results.xlsx")

    moca_cortical_md_wb = openpyxl.load_workbook("../data/BL_new_MoCA_cortical_md_results.xlsx")
    moca_lh_thickness_wb = openpyxl.load_workbook("../data/BL_new_MoCA_lh_thickness_results.xlsx")
    moca_rh_thickness_wb = openpyxl.load_workbook("../data/BL_new_MoCA_rh_thickness_results.xlsx")
    moca_lh_area_wb = openpyxl.load_workbook("../data/BL_new_MoCA_lh_area_results.xlsx")
    moca_rh_area_wb = openpyxl.load_workbook("../data/BL_new_MoCA_rh_area_results.xlsx")

    # First need all of our sheets (which contain our regional results)
    subcortical_sheets = get_sheetnames(subcortical_wb)
    rh_thickness_sheets = get_sheetnames(rh_thickness_wb)
    lh_thickness_sheets = get_sheetnames(lh_thickness_wb)
    rh_area_sheets = get_sheetnames(rh_area_wb)
    lh_area_sheets = get_sheetnames(lh_area_wb)
    cortical_md_sheets = get_sheetnames(cortical_md_wb)

    updrs_cortical_md_sheets = get_sheetnames(updrs_cortical_md_wb)
    updrs_lh_thickness_sheets = get_sheetnames(updrs_lh_thickness_wb)
    updrs_rh_thickness_sheets = get_sheetnames(updrs_rh_thickness_wb)
    updrs_lh_area_sheets = get_sheetnames(updrs_lh_area_wb)
    updrs_rh_area_sheets = get_sheetnames(updrs_rh_area_wb)

    moca_cortical_md_sheets = get_sheetnames(moca_cortical_md_wb)
    moca_lh_thickness_sheets = get_sheetnames(moca_lh_thickness_wb)
    moca_rh_thickness_sheets = get_sheetnames(moca_rh_thickness_wb)
    moca_lh_area_sheets = get_sheetnames(moca_lh_area_wb)
    moca_rh_area_sheets = get_sheetnames(moca_rh_area_wb)

    # Now we want to get the p-values
    subcortical_pvalue_dict = get_pvalues(subcortical_wb, subcortical_sheets)
    rh_thickness_pvalue_dict = get_pvalues(rh_thickness_wb, rh_thickness_sheets)
    lh_thickness_pvalue_dict = get_pvalues(lh_thickness_wb, lh_thickness_sheets)
    rh_area_pvalue_dict = get_pvalues(rh_area_wb, rh_area_sheets)
    lh_area_pvalue_dict = get_pvalues(lh_area_wb, lh_area_sheets)
    cortical_md_pvalue_dict = get_pvalues(cortical_md_wb, cortical_md_sheets)

    updrs_cortical_md_pvalue_dict = get_pvalues(updrs_cortical_md_wb, updrs_cortical_md_sheets)
    updrs_lh_thickness_pvalue_dict = get_pvalues(updrs_lh_thickness_wb, updrs_lh_thickness_sheets)
    updrs_rh_thickness_pvalue_dict = get_pvalues(updrs_rh_thickness_wb, updrs_rh_thickness_sheets)
    updrs_lh_area_pvalue_dict = get_pvalues(updrs_lh_area_wb, updrs_lh_area_sheets)
    updrs_rh_area_pvalue_dict = get_pvalues(updrs_rh_area_wb, updrs_rh_area_sheets)

    moca_cortical_md_pvalue_dict = get_pvalues(moca_cortical_md_wb, moca_cortical_md_sheets)
    moca_lh_thickness_pvalue_dict = get_pvalues(moca_lh_thickness_wb, moca_lh_thickness_sheets)
    moca_rh_thickness_pvalue_dict = get_pvalues(moca_rh_thickness_wb, moca_rh_thickness_sheets)
    moca_lh_area_pvalue_dict = get_pvalues(moca_lh_area_wb, moca_lh_area_sheets)
    moca_rh_area_pvalue_dict = get_pvalues(moca_rh_area_wb, moca_rh_area_sheets)


    # Merge hemispheric dictionaries since we want to run FDR corrections across whole brain
    rh_thickness_pvalue_dict.update(lh_thickness_pvalue_dict)
    rh_area_pvalue_dict.update(lh_area_pvalue_dict)

    updrs_rh_thickness_pvalue_dict.update(updrs_lh_thickness_pvalue_dict)
    updrs_rh_area_pvalue_dict.update(updrs_lh_area_pvalue_dict)

    moca_rh_thickness_pvalue_dict.update(moca_lh_thickness_pvalue_dict)
    moca_rh_area_pvalue_dict.update(moca_lh_area_pvalue_dict)

    # Just renaming since update() function changes stuff in place
    thickness_pvalue_dict = rh_thickness_pvalue_dict
    area_pvalue_dict = rh_area_pvalue_dict

    updrs_thickness_pvalue_dict = updrs_rh_thickness_pvalue_dict
    updrs_area_pvalue_dict = updrs_rh_area_pvalue_dict

    moca_thickness_pvalue_dict = moca_rh_thickness_pvalue_dict
    moca_area_pvalue_dict = moca_rh_area_pvalue_dict

    print(thickness_pvalue_dict)
    print(area_pvalue_dict)
    print(len(thickness_pvalue_dict))
    print(len(area_pvalue_dict))

    corr_subcortical_wb = run_fdr_correction(subcortical_pvalue_dict, subcortical_wb)
    corr_cortical_md_wb = run_fdr_correction(cortical_md_pvalue_dict, cortical_md_wb)

    corr_lh_thickness_wb, corr_rh_thickness_wb = run_fdr_correction_LR(thickness_pvalue_dict, lh_thickness_wb, rh_thickness_wb)
    corr_lh_area_wb, corr_rh_area_wb = run_fdr_correction_LR(area_pvalue_dict, lh_area_wb, rh_area_wb)

    corr_subcortical_wb.save("../data/BL_newest_corrected_subcortical_results.xlsx")
    corr_cortical_md_wb.save("../data/BL_newest_corrected_cortical_md_results.xlsx")
    #print(subcortical_pvalue_dict)

    corr_lh_thickness_wb.save("../data/BL_newest_corrected_lh_thickness_results.xlsx")
    corr_rh_thickness_wb.save("../data/BL_newest_corrected_rh_thickness_results.xlsx")

    corr_lh_area_wb.save("../data/BL_newest_corrected_lh_area_results.xlsx")
    corr_rh_area_wb.save("../data/BL_newest_corrected_rh_area_results.xlsx")

    corr_updrs_cortical_md_wb = run_updrs_fdr_correction(updrs_cortical_md_pvalue_dict, updrs_cortical_md_wb)
    
    corr_updrs_lh_thickness_wb, corr_updrs_rh_thickness_wb = run_updrs_fdr_correction_LR(updrs_thickness_pvalue_dict, updrs_lh_thickness_wb, updrs_rh_thickness_wb)
    corr_updrs_lh_area_wb, corr_updrs_rh_area_wb = run_updrs_fdr_correction_LR(updrs_area_pvalue_dict, updrs_lh_area_wb, updrs_rh_area_wb)

    corr_moca_cortical_md_wb = run_moca_fdr_correction(moca_cortical_md_pvalue_dict, moca_cortical_md_wb)

    corr_moca_lh_thickness_wb, corr_moca_rh_thickness_wb = run_moca_fdr_correction_LR(moca_thickness_pvalue_dict, moca_lh_thickness_wb, moca_rh_thickness_wb)
    corr_moca_lh_area_wb, corr_moca_rh_area_wb = run_moca_fdr_correction_LR(moca_area_pvalue_dict, moca_lh_area_wb, moca_rh_area_wb)

    corr_updrs_cortical_md_wb.save("../data/BL_corrected_new_UPDRS_cortical_md_results.xlsx")
    corr_updrs_lh_thickness_wb.save("../data/BL_corrected_new_UPDRS_lh_thickness_results.xlsx")
    corr_updrs_rh_thickness_wb.save("../data/BL_corrected_new_UPDRS_rh_thickness_results.xlsx")
    corr_updrs_lh_area_wb.save("../data/BL_corrected_new_UPDRS_lh_area_results.xlsx")
    corr_updrs_rh_area_wb.save("../data/BL_corrected_new_UPDRS_rh_area_results.xlsx")
    corr_moca_cortical_md_wb.save("../data/BL_corrected_new_MoCA_cortical_md_results.xlsx")
    corr_moca_lh_thickness_wb.save("../data/BL_corrected_new_MoCA_lh_thickness_results.xlsx")
    corr_moca_rh_thickness_wb.save("../data/BL_corrected_new_MoCA_rh_thickness_results.xlsx")
    corr_moca_lh_area_wb.save("../data/BL_corrected_new_MoCA_lh_area_results.xlsx")
    corr_moca_rh_area_wb.save("../data/BL_corrected_new_MoCA_rh_area_results.xlsx")

    return

def run_fdr_correction_LR(pvalue_dict, left_workbook, right_workbook):
 
    sex_pvalues = []
    group_pvalues = []
    educ_pvalues = []
    age_pvalues = []

    for region in pvalue_dict:
        if pvalue_dict[region]['Group'] is None: # Check for none types. Means region didn't run for whatever reason
            continue # So we just skip the region
        group_pvalues.append(pvalue_dict[region]['Group'])
        sex_pvalues.append(pvalue_dict[region]['Sex'])
        educ_pvalues.append(pvalue_dict[region]['Educ'])
        age_pvalues.append(pvalue_dict[region]['Age'])

    _, corr_group_pvalues = fdrcorrection(group_pvalues)
    _, corr_sex_pvalues = fdrcorrection(sex_pvalues)
    _, corr_educ_pvalues = fdrcorrection(educ_pvalues)
    _, corr_age_pvalues = fdrcorrection(age_pvalues)

    # Make variables for the cells so stuff is clearer later. We're gonna be updating these cells with the new (corrected) p values
    sex_pvalue_cell = "D2"
    group_pvalue_cell = "D3"
    age_pvalue_cell = "D4"
    educ_pvalue_cell = "D5"

    i=0 # Doing this instead of using enmuerate() since things don't line up exactly (more sheets than p-values)
    for region in pvalue_dict:
        if i <= 31:
            region_sheet = right_workbook[region]
            if pvalue_dict[region]['Group'] is None: # Check for none types. Means region didn't run for whatever reason
                continue # So we just skip the region
            region_sheet[group_pvalue_cell] = corr_group_pvalues[i]
            region_sheet[sex_pvalue_cell] = corr_sex_pvalues[i]
            region_sheet[educ_pvalue_cell] = corr_educ_pvalues[i]
            region_sheet[age_pvalue_cell] = corr_age_pvalues[i]
            i+=1
        else:
            region_sheet = left_workbook[region]
            if pvalue_dict[region]['Group'] is None: # Check for none types. Means region didn't run for whatever reason
                continue # So we just skip the region
            region_sheet[group_pvalue_cell] = corr_group_pvalues[i]
            region_sheet[sex_pvalue_cell] = corr_sex_pvalues[i]
            region_sheet[educ_pvalue_cell] = corr_educ_pvalues[i]
            region_sheet[age_pvalue_cell] = corr_age_pvalues[i]
            i+=1

    return left_workbook, right_workbook

def run_moca_fdr_correction(pvalue_dict, workbook):
    sex_pvalues = []
    educ_pvalues = []
    moca_pvalues = []
    age_pvalues = []

    for region in pvalue_dict:
        if pvalue_dict[region]['Sex'] is None:
            continue
        sex_pvalues.append(pvalue_dict[region]['Sex'])
        educ_pvalues.append(pvalue_dict[region]['Educ'])
        moca_pvalues.append(pvalue_dict[region]['MoCA'])
        age_pvalues.append(pvalue_dict[region]['Age'])

    _, corr_sex_pvalues = fdrcorrection(sex_pvalues)
    _, corr_educ_pvalues = fdrcorrection(educ_pvalues)
    _, corr_moca_pvalues = fdrcorrection(moca_pvalues)
    _, corr_age_pvalues = fdrcorrection(age_pvalues)

    sex_pvalue_cell = "D2"
    educ_pvalue_cell = "D3"
    moca_pvalue_cell = "D4"
    age_pvalue_cell = "D5"

    i=0
    for region in pvalue_dict:
        region_sheet = workbook[region]
        if pvalue_dict[region]['Sex'] is None:
            continue
        region_sheet[sex_pvalue_cell] = corr_sex_pvalues[i]
        region_sheet[educ_pvalue_cell] = corr_educ_pvalues[i]
        region_sheet[moca_pvalue_cell] = corr_moca_pvalues[i]
        region_sheet[age_pvalue_cell] = corr_age_pvalues[i]
        i+=1
    return workbook

def run_moca_fdr_correction_LR(pvalue_dict, left_workbook, right_workbook):
    sex_pvalues = []
    educ_pvalues = []
    moca_pvalues = []
    age_pvalues = []

    for region in pvalue_dict:
        if pvalue_dict[region]['Sex'] is None:
            continue
        sex_pvalues.append(pvalue_dict[region]['Sex'])
        educ_pvalues.append(pvalue_dict[region]['Educ'])
        moca_pvalues.append(pvalue_dict[region]['MoCA'])
        age_pvalues.append(pvalue_dict[region]['Age'])

    _, corr_sex_pvalues = fdrcorrection(sex_pvalues)
    _, corr_educ_pvalues = fdrcorrection(educ_pvalues)
    _, corr_moca_pvalues = fdrcorrection(moca_pvalues)
    _, corr_age_pvalues = fdrcorrection(age_pvalues)

    sex_pvalue_cell = "D2"
    educ_pvalue_cell = "D3"
    moca_pvalue_cell = "D4"
    age_pvalue_cell = "D5"

    i=0
    for region in pvalue_dict:
        if i <= 31:
            region_sheet = right_workbook[region]
            if pvalue_dict[region]['Sex'] is None:
                continue
            region_sheet[sex_pvalue_cell] = corr_sex_pvalues[i]
            region_sheet[educ_pvalue_cell] = corr_educ_pvalues[i]
            region_sheet[moca_pvalue_cell] = corr_moca_pvalues[i]
            region_sheet[age_pvalue_cell] = corr_age_pvalues[i]
            i += 1
        else:
            region_sheet = left_workbook[region]
            if pvalue_dict[region]['Sex'] is None:
                continue
            region_sheet[sex_pvalue_cell] = corr_sex_pvalues[i]
            region_sheet[educ_pvalue_cell] = corr_educ_pvalues[i]
            region_sheet[moca_pvalue_cell] = corr_moca_pvalues[i]
            region_sheet[age_pvalue_cell] = corr_age_pvalues[i]
            i += 1

    return left_workbook, right_workbook

def run_updrs_fdr_correction(pvalue_dict, workbook):
    sex_pvalues = []
    educ_pvalues = []
    updrs_pvalues = []
    age_pvalues = []

    for region in pvalue_dict:
        if pvalue_dict[region]['Sex'] is None:
            continue
        sex_pvalues.append(pvalue_dict[region]['Sex'])
        educ_pvalues.append(pvalue_dict[region]['Educ'])
        updrs_pvalues.append(pvalue_dict[region]['UPDRS'])
        age_pvalues.append(pvalue_dict[region]['Age'])

    _, corr_sex_pvalues = fdrcorrection(sex_pvalues)
    _, corr_educ_pvalues = fdrcorrection(educ_pvalues)
    _, corr_updrs_pvalues = fdrcorrection(updrs_pvalues)
    _, corr_age_pvalues = fdrcorrection(age_pvalues)

    sex_pvalue_cell = "D2"
    educ_pvalue_cell = "D3"
    updrs_pvalue_cell = "D4"
    age_pvalue_cell = "D5"
    visit_pvalue_cell = "D6"

    i=0
    for region in pvalue_dict:
        region_sheet = workbook[region]
        if pvalue_dict[region]['Sex'] is None:
            continue
        region_sheet[sex_pvalue_cell] = corr_sex_pvalues[i]
        region_sheet[educ_pvalue_cell] = corr_educ_pvalues[i]
        region_sheet[updrs_pvalue_cell] = corr_updrs_pvalues[i]
        region_sheet[age_pvalue_cell] = corr_age_pvalues[i]
        i+=1
    return workbook

def run_updrs_fdr_correction_LR(pvalue_dict, left_workbook, right_workbook):
    sex_pvalues = []
    educ_pvalues = []
    updrs_pvalues = []
    age_pvalues = []

    for region in pvalue_dict:
        if pvalue_dict[region]['Sex'] is None:
            continue
        sex_pvalues.append(pvalue_dict[region]['Sex'])
        educ_pvalues.append(pvalue_dict[region]['Educ'])
        if pvalue_dict[region]['UPDRS'] is not None:
            updrs_pvalues.append(pvalue_dict[region]['UPDRS'])
        age_pvalues.append(pvalue_dict[region]['Age'])

    _, corr_sex_pvalues = fdrcorrection(sex_pvalues)
    _, corr_educ_pvalues = fdrcorrection(educ_pvalues)
    _, corr_updrs_pvalues = fdrcorrection(updrs_pvalues)
    _, corr_age_pvalues = fdrcorrection(age_pvalues)

    sex_pvalue_cell = "D2"
    educ_pvalue_cell = "D3"
    updrs_pvalue_cell = "D4"
    age_pvalue_cell = "D5"

    #corr_updrs_pvalues = list(corr_updrs_pvalues)

    # Dummy value for lh_precuneus_area
    #corr_updrs_pvalues.insert(53, 0.99)

    i=0
    for region in pvalue_dict:
        if i <= 31:
            region_sheet = right_workbook[region]
            if pvalue_dict[region]['Sex'] is None:
                continue
            region_sheet[sex_pvalue_cell] = corr_sex_pvalues[i]
            region_sheet[educ_pvalue_cell] = corr_educ_pvalues[i]
            region_sheet[updrs_pvalue_cell] = corr_updrs_pvalues[i]
            region_sheet[age_pvalue_cell] = corr_age_pvalues[i]
            i += 1
        else:
            region_sheet = left_workbook[region]
            if pvalue_dict[region]['Sex'] is None:
                continue
            region_sheet[sex_pvalue_cell] = corr_sex_pvalues[i]
            region_sheet[educ_pvalue_cell] = corr_educ_pvalues[i]
            region_sheet[updrs_pvalue_cell] = corr_updrs_pvalues[i]
            region_sheet[age_pvalue_cell] = corr_age_pvalues[i]
            i += 1
    
    return left_workbook, right_workbook

def run_fdr_correction(pvalue_dict, workbook):
    # How much of my life do I spend typing? I wonder
    sex_pvalues = []
    group_pvalues = []
    educ_pvalues = []
    age_pvalues = []

    for region in pvalue_dict:
        if pvalue_dict[region]['Group'] is None: # Check for none types. Means region didn't run for whatever reason
            continue # So we just skip the region
        group_pvalues.append(pvalue_dict[region]['Group'])
        sex_pvalues.append(pvalue_dict[region]['Sex'])
        educ_pvalues.append(pvalue_dict[region]['Educ'])
        age_pvalues.append(pvalue_dict[region]['Age'])

    _, corr_group_pvalues = fdrcorrection(group_pvalues)
    _, corr_sex_pvalues = fdrcorrection(sex_pvalues)
    _, corr_educ_pvalues = fdrcorrection(educ_pvalues)
    _, corr_age_pvalues = fdrcorrection(age_pvalues)

    # Make variables for the cells so stuff is clearer later. We're gonna be updating these cells with the new (corrected) p values

    sex_pvalue_cell = "D2"
    group_pvalue_cell = "D3"
    age_pvalue_cell = "D4"
    educ_pvalue_cell = "D5"

    i=0 # Doing this instead of using enmuerate() since things don't line up exactly (more sheets than p-values)
    for region in pvalue_dict:
        region_sheet = workbook[region]
        if pvalue_dict[region]['Group'] is None: # Check for none types. Means region didn't run for whatever reason
            continue # So we just skip the region
        region_sheet[group_pvalue_cell] = corr_group_pvalues[i]
        region_sheet[sex_pvalue_cell] = corr_sex_pvalues[i]
        region_sheet[educ_pvalue_cell] = corr_educ_pvalues[i]
        region_sheet[age_pvalue_cell] = corr_age_pvalues[i]
        i+=1

    return workbook

def get_pvalues(workbook, sheet_names):
    # Yet another dictionary of dictionaries
    # This one will map regions to dictionaries which map parameters to their p-values for that region

    pvalue_dict = {}

    for sheet_name in sheet_names:
        worksheet = workbook[sheet_name]
        rows = list(worksheet.rows)
        for row in rows:
            parameter = row[1].value
            p_value = row[3].value

            if parameter == "Parameter": # Don't want this row
                continue

            if sheet_name in pvalue_dict:
                pvalue_dict[sheet_name][parameter] = p_value # Sheet name is region name
            else:
                pvalue_dict[sheet_name] = {parameter: p_value}
    return pvalue_dict

def get_sheetnames(workbook):
    sheetname_list = []

    for sheetname in workbook.sheetnames:
        if sheetname == "Sheet": # Dummy sheet, probably could have fixed this in MLM script but... lazy...
            continue
        sheetname_list.append(sheetname)

    return sheetname_list

if __name__ == "__main__":
    main()