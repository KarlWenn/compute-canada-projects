import pandas as pd
import numpy as np
import os
import openpyxl
from openpyxl import load_workbook

demographics_csv = "../data/ppmi_mri_matched_data_karl.csv"

def main():
    sub_ses_list, age_dict, sex_dict, educ_dict, diagnosis_dict, updrs_dict, moca_dict = make_demographic_dicts()

    sessions = ["All", "ses-BL", "ses-V04", "ses-V06", "ses-V08", "ses-V10"]

    wb = "../data/fw_demographics_final.xlsx"

    dwi_df = pd.read_csv("../data/dwi_demographics_v4.csv")
    
    for session in sessions:
        make_categorical_demographic_sheet(dwi_df, wb, session, sub_ses_list, age_dict, sex_dict, educ_dict, diagnosis_dict, updrs_dict, moca_dict)

def make_categorical_demographic_sheet(df, wb, ses, sub_ses_list, age_dict, sex_dict, educ_dict, diagnosis_dict, updrs_dict, moca_dict):
    num_sub_list = []
    average_age_list = []
    num_male_list = []
    num_female_list = []
    average_educ_list = []
    average_moca_list = []
    average_updrs_list = []

    categories = ['ALL1', 'APPA1', 'LRRL1', 'SD1']
    for category in categories:
        # For each category get the demographic measure needed to fill out the dictionary below
        average_age_overall, number_male_overall, number_female_overall, average_educ_overall, number_spd_overall, number_hc_overall, average_updrs_overall, average_moca_overall, \
        average_age_spd, number_male_spd, number_female_spd, average_educ_spd, average_updrs_spd, average_moca_spd, \
        average_age_hc, number_male_hc, number_female_hc, average_educ_hc, average_updrs_hc, average_moca_hc = get_demographics(df, category, ses, sub_ses_list, age_dict, sex_dict, educ_dict, diagnosis_dict, updrs_dict, moca_dict)

        # Append everything to lists, making sure to keep the proper order, again corresponding to the order of things in the dictionary below
        # I.e. Combined measures, then healthy controls, then sPD patients
        num_sub_list.append(number_hc_overall+number_spd_overall)
        num_sub_list.append(number_hc_overall)
        num_sub_list.append(number_spd_overall)

        average_age_list.append(average_age_overall)
        average_age_list.append(average_age_hc)
        average_age_list.append(average_age_spd)

        num_male_list.append(number_male_overall)
        num_male_list.append(number_male_hc)
        num_male_list.append(number_male_spd)

        num_female_list.append(number_female_overall)
        num_female_list.append(number_female_hc)
        num_female_list.append(number_female_spd)

        average_educ_list.append(average_educ_overall)
        average_educ_list.append(average_educ_hc)
        average_educ_list.append(average_educ_spd)

        average_moca_list.append(average_moca_overall)
        average_moca_list.append(average_moca_hc)
        average_moca_list.append(average_moca_spd)

        average_updrs_list.append(average_updrs_overall)
        average_updrs_list.append(average_updrs_hc)
        average_updrs_list.append(average_updrs_spd)

    demographic_dict = {
        "Category": ['ALL1', 'ALL1', 'ALL1', 'APPA1', 'APPA1', 'APPA1', 'LRRL1', 'LRRL1', 'LRRL1', 'SD1', 'SD1', 'SD1'],
        "Diagnosis": ['Combined', 'HC', 'sPD', 'Combined', 'HC', 'sPD', 'Combined', 'HC', 'sPD', 'Combined', 'HC', 'sPD'],
        "No. Subjects": num_sub_list,
        "Average Age (years)": average_age_list,
        "No. Male Subjects": num_male_list,
        "No. Female Subjects": num_female_list,
        "Average Education (years)": average_educ_list,
        "Average MoCA": average_moca_list,
        "Average UPDRS": average_updrs_list
    }

    data_df = pd.DataFrame(demographic_dict)

    if os.path.isfile(wb):
        results_wb = load_workbook(wb)
    else:
        results_wb = create_excel_wb(wb)

    if ses in results_wb.sheetnames:
        with pd.ExcelWriter(wb, mode='a', if_sheet_exists='replace') as writer:
            data_df.to_excel(writer, sheet_name=ses)
    else:
        if os.path.isfile(wb):
            with pd.ExcelWriter(wb, mode='a') as writer:
                data_df.to_excel(writer, sheet_name=ses)
        else:
            with pd.ExcelWriter(wb, mode='w') as writer:
                data_df.to_excel(writer, sheet_name=ses)

    return 

def get_demographics(df, category, ses, sub_ses_list, age_dict, sex_dict, educ_dict, diagnosis_dict, updrs_dict, moca_dict):
    age_values = []
    sex_values = []
    educ_values = []
    diagnosis_values = []
    updrs_values = []
    moca_values = []

    age_spd_values = []
    sex_spd_values = []
    educ_spd_values = []
    updrs_spd_values = []
    moca_spd_values = []

    age_hc_values= []
    sex_hc_values = []
    educ_hc_values = []
    updrs_hc_values = []
    moca_hc_values = []

    for index, row in df.iterrows():
        subj_category = row['Category']
        # Latter part of this disjunction takes care of the fact none of the subjects have ALL1 has their category
        if subj_category == category or category == "ALL1":
            subject = row['Subject']
            session = row['Session']
            print(subject, session)
            # First check if the subject session is in the demographic CSV at all, and continue if not
            # No reason to go forward, that is
            if (subject, session) not in sub_ses_list:
                continue
            # Also have no interest in prod subjects as far as demographics are concerned
            if diagnosis_dict[(subject, session)] == 'prod':
                continue
            if ses == "ses-BL":
                if session != "ses-BL":
                    continue
            elif ses == "ses-V04":
                if session != "ses-V04":
                    continue
            elif ses == "ses-V06":
                if session != "ses-V06":
                    continue
            elif ses == "ses-V08":
                if session != "ses-V08":
                    continue
            elif ses == "ses-V10":
                if session != "ses-V10":
                    continue
            try:
                age_values.append(age_dict[(subject, session)])
            except KeyError:
                pass
            sex = sex_dict[(subject, session)]
            education = educ_dict[(subject, session)]

            sex_values.append(sex)
            educ_values.append(education)
            
            try:
                diagnosis_values.append(diagnosis_dict[(subject, session)])
            except KeyError:
                continue

            try:
                updrs_values.append(updrs_dict[(subject, session)])
            except KeyError:
                pass
            try:
                moca_values.append(moca_dict[(subject, session)])
            except KeyError:
                pass
            if diagnosis_dict[(subject, session)] == "ctrl":
                age_hc_values.append(age_dict[(subject, session)])
                sex_hc_values.append(sex_dict[(subject, session)])
                educ_hc_values.append(educ_dict[(subject, session)])
                try:
                    updrs_hc_values.append(updrs_dict[(subject, session)])
                except KeyError:
                    pass
                try:
                    moca_hc_values.append(moca_dict[(subject, session)])
                except KeyError:
                    pass
            elif diagnosis_dict[(subject, session)] == "park":
                try: # Park subjects are missing age values sometimes, can just pass over them. sub-4061 ses-BL is one such example
                    age_spd_values.append(age_dict[(subject, session)])
                except KeyError:
                    pass
                sex_spd_values.append(sex_dict[(subject, session)])
                educ_spd_values.append(educ_dict[(subject, session)])
                try:
                    updrs_spd_values.append(updrs_dict[(subject, session)])
                except KeyError:
                    pass
                try:
                    moca_spd_values.append(moca_dict[(subject, session)])
                except KeyError:
                    pass 
    print(category, ses)
    print(age_values)
    average_age_overall = sum(age_values)/len(age_values)
    number_male_overall = sex_values.count('male')
    number_female_overall = sex_values.count('female')
    average_educ_overall = sum(educ_values)/len(educ_values)
    number_spd_overall = diagnosis_values.count('park')
    number_hc_overall = diagnosis_values.count('ctrl')
    average_updrs_overall = sum(updrs_values)/len(updrs_values)
    average_moca_overall = sum(moca_values)/len(moca_values)

    average_age_spd = sum(age_spd_values)/len(age_spd_values)
    number_male_spd = sex_spd_values.count('male')
    number_female_spd = sex_spd_values.count('female')
    average_educ_spd = sum(educ_spd_values)/len(educ_spd_values)
    average_updrs_spd = sum(updrs_spd_values)/len(updrs_spd_values)
    average_moca_spd = sum(moca_spd_values)/len(moca_spd_values)

    average_age_hc = sum(age_hc_values)/len(age_hc_values)
    number_male_hc = sex_hc_values.count('male')
    number_female_hc = sex_hc_values.count('female')
    average_educ_hc = sum(educ_hc_values)/len(educ_hc_values)
    average_updrs_hc = sum(updrs_hc_values)/len(updrs_hc_values)
    average_moca_hc = sum(moca_hc_values)/len(moca_hc_values)
    
    return average_age_overall, number_male_overall, number_female_overall, average_educ_overall, number_spd_overall, number_hc_overall, average_updrs_overall, average_moca_overall, \
    average_age_spd, number_male_spd, number_female_spd, average_educ_spd, average_updrs_spd, average_moca_spd, \
    average_age_hc, number_male_hc, number_female_hc, average_educ_hc, average_updrs_hc, average_moca_hc

def make_demographic_dicts():
    age_dict = {}
    sex_dict = {}
    educ_dict = {}
    diagnosis_dict = {}
    updrs_dict = {}
    moca_dict = {}
    sub_ses_list = []

    demographic_df = pd.read_csv(demographics_csv)

    for index, row in demographic_df.iterrows():
        subject = f"sub-{row['participant_id']}"
        session = f"ses-{row['visit']}"
        sex = row['sex']
        age = row['age']
        education = row['education']
        diagnosis = row['diagnosis']
        updrs = row['updrs3']
        moca = row['moca']

        sub_ses_list.append((subject, session))

        # Check if values are missing, skip them if so.
        # Will still take the other values if they are available.
        if not np.isnan(age):  # Accounting for potential missing data, here and below
            age_dict[(subject, session)] = age
        if sex != np.nan:
            sex_dict[(subject, session)] = sex
        if not np.isnan(education):
            educ_dict[(subject, session)] = education
        if diagnosis != np.nan:
            diagnosis_dict[(subject, session)] = diagnosis

        if not np.isnan(updrs):
            updrs_dict[(subject, session)] = updrs

        if not np.isnan(moca):
            moca_dict[(subject, session)] = moca

    return sub_ses_list, age_dict, sex_dict, educ_dict, diagnosis_dict, updrs_dict, moca_dict

def create_excel_wb(filepath):
    wb = openpyxl.Workbook()
    wb.save(filepath)
    return wb

if __name__ == "__main__":
    main()